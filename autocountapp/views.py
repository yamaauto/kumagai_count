from flask import Flask, render_template, request, url_for, redirect, flash, session, jsonify
from flask_apscheduler import APScheduler
from autocountapp import app
from autocountapp import handle_ss, handle_plc
from werkzeug.utils import secure_filename

import os
import sys
import datetime
import csv
import io
import openpyxl
import time


# UPLOAD_FOLDER = 'uploads'
# app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx'} # 許可するファイル拡張子

# スケジューラー設定クラス
class Config:
    SCHEDULER_API_ENABLED = True  # API有効化（任意）

app.config.from_object(Config())

scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()

# 定期実行した内容を保存する変数
job_output = ""
time_count = 0

# 定期実行したい関数
# @scheduler.task('interval', id='job_1', seconds=300)
# def job():
#     global job_output
#     global time_count
#     time_count = time_count + 1
#     job_output = "10秒ごとに実行中！"
#     print(time_count)


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# # 毎日12時にその時点までのプレスログ（ローカルに保存しているもののみ）を昼プレスログシートに記入する
# @scheduler.task('cron', id='daily_job', hour=12, minute=00)
# def daily_task():
#     print("毎日12時の定期タスク！")
#     machine_datas = get_machine_status()
#     count = 3000
#     # count = handle_plc.get_count()
#     for row in machine_datas:
#         if row[1]['status'] == "稼働中":
#             new_row_ss = [row[1]['ope_id'], row[1]['date_st'], row[0], row[1]['item'], row[1]['time_st'], "", row[1]['count_st'], count, 0]
#             handle_ss.add_new_row(new_row_ss, "プレスログ")


machine_order = ['A-500', 'P-300', 'S-250', 'R-150', 'K-150', 'J-300', 'E-150', 'M-45']

#生産画面
@app.route('/', methods=['GET', 'POST'])
def autocount_home():
    dt_now = datetime.datetime.now()
    if request.method == 'GET':
        press = request.args.get('press')
    elif request.method == 'POST':
        # press = request.form['press']
        press = ""
        ope_val_ = {'date':dt_now.strftime('%Y/%m/%d')}
        for key, value in request.form.items():
            ope_val_[key] = value
        a = regist_status(ope_val_)
        flash('登録に成功しました!')
        return redirect(url_for('autocount_home', press=press))

    machine_datas = get_machine_status()
    for machine in machine_datas:
        if machine[1]["status"] == "停止中":
            machine[1]['item_name'] = ""
        else:
            search_items_result = search_items(machine[1]["item"])
            if len(search_items_result) > 4:
                machine[1]['item_name'] = search_items_result[4]
            else:
                machine[1]['item_name'] = ""

    global machine_order
    order_map = {key: i for i, key in enumerate(machine_order)}
    machine_datas = sorted(machine_datas, key=lambda item: order_map[item[0]])

    status_list = [[i[0],i[1]["status"]] for i in machine_datas]
    items = get_items()
    return render_template('tablet.html', machine_datas=machine_datas, items=items, status_list=status_list, press=press)



#調整記入
@app.route('/fixmachine', methods=['GET', 'POST'])
def fix_machine():
    if request.method == 'GET':
        return redirect(url_for('autocount_home'))
    elif request.method == 'POST':
        press = request.form['press']
        machine = request.form['machine']
        check_type = request.form['check_type']
        machine_data = get_eachmachine_status(machine)
        message = machine_data[1]['message']
        status = machine_data[1]['status']
        if check_type == 'setup': #検査の場合、選択された商品・新規ID・を返す　ローカルで更新するのはＩＤと稼働状況のみ
            try:
                if request.form['return_btn']=="home": #戻る処理
                    flash('処理を中止しました')
                    return redirect(url_for('autocount_home', press=press))
            except:
                pass
            if status == "停止中":
                item = request.form['item'] #段取り開始時のみ商品の情報を持たないのでフォームから取得
                btn_type = 'start'
            elif status == "段取り一時停止中":
                item = machine_data[1]['item']
                btn_type = 'start'
            elif status == "段取り中":
                item = machine_data[1]['item']
                btn_type = 'end'
        elif check_type == 'fix': #調整の場合、既にローカルCSVに記入されているID・商品を返す
            item = machine_data[1]["item"]
            try:
                if request.form['return_btn']=="home": #戻る処理
                    flash('処理を中止しました')
                    return redirect(url_for('autocount_home'))
            except:
                pass
            if status =="稼働中" or status == "調整一時停止中":
                btn_type = 'start'
            elif status =="調整中":
                btn_type = 'end'
        elif check_type == 'check': #検査の場合、選択された商品・新規ID・を返す　ローカルで更新するのはＩＤと稼働状況のみ
            try:
                if request.form['return_btn']=="home": #戻る処理
                    flash('処理を中止しました')
                    return redirect(url_for('autocount_home'))
            except:
                pass
            if status =="停止中":
                item = request.form['item'] #検査開始時のみ商品の情報を持たないのでフォームから取得
                btn_type = 'start'
            elif status =="段取り済" or status == "検査一時停止中":
                item = machine_data[1]['item']
                btn_type = 'start'
            elif status =="検査中":
                item = machine_data[1]['item']
                btn_type = 'end'
        item_data = search_items(item)
        return render_template('fixmachine.html', machine=machine, item=item, item_data=item_data, check_type=check_type, btn_type=btn_type, press=press, message=message)



#調整登録処理
@app.route('/fixmachineregist', methods=['GET', 'POST'])
def fix_machine_regist():
    if request.method == 'POST':
        press = request.form['press']
        dt_now = datetime.datetime.now()
        check_date = dt_now.strftime('%Y/%m/%d')
        machine = request.form['machine']
        check_type = request.form['check_type']
        btn_type = request.form['btn_type']

        machine_data = get_eachmachine_status(machine)
        message = machine_data[1]['message']
        if check_type == 'check' or check_type == 'setup': #検査 or 段取り
            check_type_ja = "検査" if check_type=="check" else "段取り"
            if btn_type == 'start': #ここで新規ID発行　/fixmachineではIDをローカルのまま流しているだけなので検査開始時点では値なし
                machine_status = check_type_ja + "中"
                item = request.form['item']
                check_st_time = dt_now.strftime('%H:%M')
                time_st = ""
                date_st = ""
                ope_id = check_machine_id(machine)
                count=""
                message = ""
                message_flg()
                flash(f"{machine} の{check_type_ja}を開始しました")
            elif btn_type == 'end':
                machine_status = "稼働中" if check_type=="check" else "段取り済" #検査終了の場合はそのまま検査へ段取り終了の場合は段取り済に
                item = machine_data[1]['item']
                check_st_time = machine_data[1]["fixcheck_st_time"] #検査開始時刻
                if check_type=="check":
                    time_st = machine_data[1]['time_st'] if machine_data[1]['time_st'] else dt_now.strftime('%H:%M') #生産開始時刻
                    date_st = machine_data[1]['date_st'] if machine_data[1]['date_st'] else dt_now.strftime('%Y/%m/%d')
                    count = machine_data[1]['count_st']if machine_data[1]['count_st'] else handle_plc.get_count(machine)
                    # count = handle_plc.get_count(machine) if check_type=="check" else 0
                if check_type=='setup':
                    time_st = ""
                    date_st = ""
                    count = ""
                    message = ""
                    message_flg()
                ope_id = machine_data[1]['ope_id']
                comment = request.form['comment']

                #スプレッドシート更新
                new_row_ss = [ope_id, check_date, machine, item, check_st_time, dt_now.strftime('%H:%M'), '', "検査" if check_type=="check" else "段取り", comment]
                handle_ss.add_new_row(new_row_ss, "調整記録")
                check_st_time = "" #ローカルの更新用に空欄にしておく
                flash(f"{machine} の{check_type_ja}を終了しました")
            update_machine_status([ope_id,machine_status,date_st,time_st,count,item,check_st_time,message], machine)
        elif check_type == 'fix': #調整
            ope_id = machine_data[1]['ope_id']
            item = machine_data[1]['item']
            if btn_type == 'start': #稼働状況を調整中に変更してローカルの各機械のCSV末尾に調整開始時刻を追記する
                machine_status = "調整中"
                check_st_time = dt_now.strftime('%H:%M')
                message = ""
                message_flg()
                flash(f"{machine} の調整を開始しました")
            elif btn_type == 'end': #ここで調整開始時刻をＣＳＶから取得する
                machine_status = "検査中"
                check_st_time = machine_data[1]["fixcheck_st_time"]
                comment = request.form['comment']
                #スプレッドシート更新
                new_row_ss = [ope_id, check_date, machine, item, check_st_time, dt_now.strftime('%H:%M'), '', "調整", comment] #スプレッドシートに調整結果を記録
                handle_ss.add_new_row(new_row_ss, "調整記録")
                check_st_time = dt_now.strftime('%H:%M') #検査の開始時刻
                message = ""
                message_flg()
                flash(f"{machine} の調整を終了し検査を開始しました")                
            update_machine_status([ope_id,machine_status,machine_data[1]['date_st'],machine_data[1]['time_st'],machine_data[1]['count_st'],item,check_st_time,message], machine)

        
        return redirect(url_for('autocount_home', press=press))

#品質保証部メッセージ
@app.route('/qa', methods=['GET', 'POST'])
def qa():
    if request.method == 'GET':
        True
    elif request.method == 'POST':
        # プレス機ごとのコメント更新
        add_message(request)
        # アップロードされたCSVの読み取り、更新
        file = request.files['file']
        if file:
            update_item_xlsx(file)
        else:
            flash('メッセージを更新しました')
    global machine_order
    machine_datas = get_machine_status()
    order_map = {key: i for i, key in enumerate(machine_order)}
    machine_datas = sorted(machine_datas, key=lambda item: order_map[item[0]])

    val_items = get_items()
    return render_template('management.html', machine_datas=machine_datas, val_items=val_items)



#パス登録（いらないかも
@app.route('/settingpath', methods=['GET', 'POST'])
def setting_path():
    return render_template('settingpath.html')



#新規機械登録
@app.route('/settingmachine', methods=['GET', 'POST'])
def setting_machine():
    res = ''
    if request.method == 'POST':
        new_machine = request.form['new_machine']
        if new_machine == '':
            res = '名称を記入してください'
        else:
            res = add_machine_data(new_machine)
            res = '登録成功' if res == True else '既に登録されています'
    datas = get_machine_status()
    return render_template('settingmachine.html', datas=datas, res=res)


@app.route('/messages_history', methods=['GET', 'POST'])
def messages_history():
    if request.method == 'GET':
        rows = reversed(get_row_messages())
    return render_template('messages_history.html', rows=rows)


@app.route('/getmessage', methods=['GET'])
def get_comment():
        try:
            rows = get_row_messages()
            late_message = rows[-1]
            if int(late_message[3]) == 0 and late_message[2] != "":
                text_ = "プレス機:"+late_message[1]+", "+late_message[2]
            elif int(late_message[3]) == 1:
                text_ = ""
            return jsonify(text=text_)
        except FileNotFoundError:
            # ファイルが見つからない場合はエラーメッセージを返す
            return jsonify(text="エラー: data.txtが見つかりません。"), 404

##################################################

def regist_status(ope_val):
    dt_now = datetime.datetime.now()
    date_today = dt_now.date()
    global machine_order
    if ope_val['operation'] == 'allnxtday':
        ope_val['machine'] = machine_order
    else:
        ope_val['machine'] = [ope_val['machine']]

    datas = get_machine_status()
    for machine in ope_val['machine']:
        for i in range(len(datas)): #稼働状態シートから、どのプレス機か判定, iをその後使いまわす
            if machine in datas[i]:
                break
        # machine = ope_val['machine']
        status = datas[i][1]['status']
        message = datas[i][1]['comment']
        ope_id = datas[i][1]['ope_id'] 
        item = datas[i][1]['item']
        check_st_time = datas[i][1]["fixcheck_st_time"]
        if ope_val['operation'] == 'start': #生産開始
            ope_id = check_machine_id(machine) #startの処理は一時停止している日を跨いだ加工の再開のみなので
            item = get_eachmachine_status(machine)[1]['item']
            # count = 1000
            count = handle_plc.get_count(machine)
            new_row = [ope_id, "稼働中", dt_now.strftime('%Y/%m/%d'), dt_now.strftime('%H:%M'), count, item, ""]
            update_machine_status(new_row, machine)
        elif ope_val['operation'] == 'end' or ope_val['operation'] == 'nxtday' or ope_val['operation'] == 'allnxtday': #生産終了
            if datas[i][1]["date_st"]:
                count = handle_plc.get_count(machine)
                date_st = datetime.datetime.strptime(datas[i][1]["date_st"], "%Y/%m/%d").date()
                if date_today > date_st: #稼働開始が前日以前だった場合の処理 不要かも
                    delta_days = (date_today - date_st).days
                    time_ed = str(int(dt_now.strftime('%H'))+int(delta_days)*24) + ":" + str(dt_now.strftime('%M'))
                else:
                    time_ed = dt_now.strftime('%H:%M')
                time_st = datas[i][1]["time_st"]
                count_st = int(datas[i][1]["count_st"])
                count_ed = count

                if status == "稼働中":
                    if ope_val['operation'] == 'end':
                        new_row = ["", "停止中", "", "", "", "", ""]
                        flg_nxtday = 0
                    elif ope_val['operation'] == 'nxtday' or ope_val['operation'] == 'allnxtday':
                        new_row = [ope_id, "稼働一時停止中", "", "", "", item, ""]
                        flg_nxtday = 1
                    # スプレッドシート更新
                    # log_id = handle_ss.get_a_col("プレスログ", 1)
                    # datas = handle_ss.get_all_row("プレスログ")

                    # ローカル更新
                    # new_row = ["", "停止中", "", "", "", ""]
                    update_machine_status(new_row, machine)

                #稼働開始日がある場合の調整・検査は、日跨ぎ処理をする場合に生産停止処理も同時に行う必要がある
                if status == "調整中" or status == "検査中":
                    if status == "調整中":
                        machine_status = "調整一時停止中"
                        check_type_ja = "調整"
                        flg_nxtday = 1
                    elif status == "検査中":
                        machine_status = "検査一時停止中"
                        check_type_ja = "検査"
                        flg_nxtday = 1
                    update_machine_status([ope_id,machine_status,"","","",item,"",message], machine)
                    check_date = dt_now.strftime('%Y/%m/%d')
                    new_row_ss = [ope_id, check_date, machine, item, check_st_time, dt_now.strftime('%H:%M'), '', check_type_ja, "日跨ぎ一時停止"]
                    handle_ss.add_new_row(new_row_ss, "調整記録")
                
                new_row_ss = [ope_id, date_today.strftime('%Y/%m/%d'), machine, item, time_st, time_ed, count_st, 0, count_ed, "", "", flg_nxtday]
                handle_ss.add_new_row(new_row_ss, "プレスログ")

            elif status == "段取り中" or status == "検査中": #ここから調整項目の日跨ぎ処理
                if status == "段取り中":
                    machine_status = "段取り一時停止中"
                    check_type_ja = "段取り"
                elif status == "検査中":
                    machine_status = "検査一時停止中"
                    check_type_ja = "検査"
                # elif status == "調整中":
                #     machine_status = "調整一時停止中"
                #     check_type_ja = "調整"
                update_machine_status([ope_id,machine_status,"","","",item,"",message], machine)
                check_date = dt_now.strftime('%Y/%m/%d')
                new_row_ss = [ope_id, check_date, machine, item, check_st_time, dt_now.strftime('%H:%M'), '', check_type_ja, "日跨ぎ一時停止"]
                handle_ss.add_new_row(new_row_ss, "調整記録")
            # else:
            #     return False
        # elif ope_val['operation'] == 'endday': #本日生産終了
        
    


    return True



def add_machine_data(new_machine_name):
    csv_path = resource_path('autocountapp/machines/' + new_machine_name + '.csv')
    try:
        # os.makedirs(dir_path+new_machine_name)
        # handle_ss.add_new_row([new_machine_name,"停止中"], "稼働状態")
        with open(csv_path, mode='w', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["","停止中","","","","","",""])
        return True
    except Exception as e:
        print("エラーが発生しました:", e)
        return False
    


def add_message(request): #管理画面からコメントと商品を更新
    dt_now = datetime.datetime.now()
    dt_now = dt_now.strftime('%Y/%m/%d %H:%M:%S')
    
    for key,value in request.form.items():
        # if value :
            if key == "file":
                pass
            else:
                csv_path = resource_path('autocountapp/machines/' + key + '.csv')
                with open(csv_path, mode='r', encoding='utf-8') as f:
                    val = csv.reader(f)
                    for row in val:
                        a = row
                        if a[7] == value:
                            break
                        else:    
                            a[7] = value
                            with open(resource_path('autocountapp/messages_history.csv'), mode='a', encoding='utf-8') as f2:
                                    writer = csv.writer(f2,  lineterminator="\n")
                                    writer.writerow([dt_now, key, value, 0])
                                    break
                with open(csv_path, mode='w', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(a)
    


def get_items():
    csv_path = resource_path('autocountapp/items.csv')
    items_data = []
    with open(csv_path, mode='r', encoding='utf-8') as f:
        val = csv.reader(f)
        for row in val:
            items_data.append(row)
    return items_data

def search_items(sku):
    csv_path = resource_path('autocountapp/items.csv')
    matched_row = []
    with open(csv_path, mode='r', encoding='utf-8') as f:
        rows = list(csv.reader(f))
        for row in reversed(rows):
            if str(row[3]) == str(sku):
                matched_row = row
                break
    return matched_row


def get_eachmachine_status(machine_name): #機械ごとのCSVを読み込み、 [名前：{status:稼働状態, date_st:開始日, time_st:開始時間, count_st:開始カウント, item:商品名, comment:コメント}]
    csv_path = resource_path('autocountapp/machines/' + machine_name + ".csv" )
    machine_data = []
    try:
        with open(csv_path, mode='r', encoding='utf-8') as f:
            val = csv.reader(f)
            for row in val:
                machine_data = [machine_name,{"ope_id":row[0], "status":row[1], "date_st":row[2], "time_st":row[3], "count_st":row[4], "item":row[5], "fixcheck_st_time":row[6], "message":row[7]}]
    except: #存在しない時の処理作る
        val = ''
    return machine_data


def get_machine_status(): #機械ごとのCSVを読み込み、 [名前：{status:稼働状態, date_st:開始日, time_st:開始時間, count_st:開始カウント, item:商品名, comment:コメント}]
    # dir_path = 'autocountapp/machines/'
    dir_path = resource_path('autocountapp/machines/')
    machine_names = sorted(os.listdir(dir_path))
    machine_datas = []
    for machine_name in machine_names:
        csv_path = dir_path + machine_name #各CSVのパス     
        try:
            with open(csv_path, mode='r', encoding='utf-8') as f:
                val = csv.reader(f)
                for row in val:
                    idx = machine_name.find('.csv')
                    machine_datas.append([machine_name[:idx],{"ope_id":row[0], "status":row[1], "date_st":row[2], "time_st":row[3], "count_st":row[4], "item":row[5], "fixcheck_st_time":row[6], "comment":row[7], "item_name":""}])
        except: #存在しない時の処理作る
            val = ''
    return machine_datas

def update_machine_status(new_row, machine_name):
    csv_path = resource_path('autocountapp/machines/'+ machine_name + '.csv')
    with open(csv_path, mode='r', encoding='utf-8') as f:
        val = csv.reader(f)
        for row in val:
            a = row
            break
    if len(new_row) == 7:#コメントをそのまま追加   
        new_row.append(a[7])
    with open(csv_path, mode='w', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(new_row)


#for exe path
def resource_path(relative_path):
    # if hasattr(sys, '_MEIPASS'):
    #     base_path = sys._MEIPASS
    # else:
    #     base_path = os.path.abspath(".")
    if getattr(sys, 'frozen', False):  # exe 化されているとき
        base_path = os.path.dirname(sys.executable)
    else:  # Python スクリプト実行時
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# 生産開始、検査終了のタイミングでＩＤのチェック、更新を行う
def check_id_regist_today():
    dt_now = datetime.datetime.now()
    date_today = dt_now.date()
    csv_path = resource_path('autocountapp/id_regist.csv')
    d = []
    with open(csv_path, mode='r', encoding='utf-8') as f:
        val = csv.reader(f)
        for row in val:
            d.append(row)

    if datetime.datetime.strptime(d[0][0], "%Y/%m/%d").date() < date_today:
        id_num = 0
    else:
        id_num = int(d[0][1]) + 1

    with open(csv_path, mode='w', encoding='utf-8') as f:
        writer = csv.writer(f)  
        writer.writerow([dt_now.strftime('%Y/%m/%d'), id_num])

    return id_num


# プレス機ごとに前日から持ち越しているIDががないかをチェックする。ある場合はIDを返しない場合はFalseを返す
def check_machine_id(machine):
    dt_now = datetime.datetime.now()
    csv_path = resource_path('autocountapp/machines/'+machine+'.csv')
    d = []
    with open(csv_path, mode='r', encoding='utf-8') as f:
        val = csv.reader(f)
        for row in val:
            d.append(row)

    if d[0][0]: #既にIDがある場合(日を跨いでいる場合、もしくは段取り後の検査)はそれを返す
        return d[0][0]
    else: #IDが設定されていない新規の場合は作成
        ope_id_num = check_id_regist_today()
        ope_id = dt_now.strftime('%Y%m%d') + str(ope_id_num).zfill(3)
        return ope_id   
    
def update_item_xlsx(file):
    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            
            # アップロードされたファイルのバイトストリームを直接読み込む
            # これにより、ファイルをディスクに保存せずに直接openpyxlで扱えます
            # openpyxlはファイルオブジェクトを直接扱えます
            wb = openpyxl.load_workbook(file.stream, data_only=True)
            
            datas = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                print(f'--- {sheet} ---')
                
                customer_no = None
                customer_name = None

                for i, row in enumerate(ws.iter_rows(values_only=True, min_row=2)):
                    if i == 0:
                        customer_no = row[0]
                        customer_name = row[1]
                    elif i == 1:
                        continue # ヘッダー行をスキップ
                    else:
                        if row[0] is None:
                            break
                        datas.append([
                            customer_no, 
                            customer_name, 
                            row[0], 
                            row[1], 
                            row[2], 
                            row[3], 
                            row[4], 
                            row[5], 
                            row[6], 
                            row[7]
                        ])
            
            # CSVファイルへの書き込み
            csv_path = resource_path('autocountapp/items.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(datas) # writerowをループする代わりにwriterowsを使う
            
            # スプレッドシートの更新
            handle_ss.update_item_ss(datas)
            
            flash(f'ファイル "{filename}" をアップロードし、商品を更新しました。', 'success')
            
        except Exception as e:
            flash(f'ファイル処理中にエラーが発生しました: {e}', 'error')
            print(f"ファイル処理エラー: {e}")
            
    else:
        flash('許可されていないファイルタイプです。XLSXファイルのみアップロードできます。', 'error')



def get_row_messages():
    try:
        csv_path = resource_path('autocountapp/messages_history.csv')
        with open(csv_path, mode='r', encoding='utf-8') as f:
            rows = list(csv.reader(f))
    except Exception as e:
        rows = []
        print(f"CSV読み込みエラー: {e}")

    return rows

def message_flg():
    csv_path = resource_path('autocountapp/messages_history.csv')
    all_rows = []
    with open(csv_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        all_rows = list(reader) # 残りのデータ行をリスト化
        for i, row in reversed(list(enumerate(all_rows))):
            machine = row[1]
            status = row[3]
            if status == "0":
                all_rows[i][3] = 1
                break
            elif status == "1":
                break
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(all_rows) # 更新されたデータ行を全て書き込む


    # for row in lst:
    # # indexが2の行を更新したい
    # if row[0] == '2':
    #     row[0], row[1] = a[0], a[1]